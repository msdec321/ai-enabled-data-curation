#!/usr/bin/env python3
"""Build the PCORnet CDM valueset catalog the agent uses for conformance checks.

Two authoritative sources, merged into one catalog keyed by TABLE.COLUMN:

  1. docs/PCORI_CDM_SPECIFICATION.pdf      — the CDM spec. Categorical fields carry
     an inline "Predefined Value Sets" column (SEX, RACE, DX_TYPE, ...). Small,
     enumerable, embedded directly under each column.
  2. docs/PCORNET_ADDITIONAL_VALUESETS.xlsx — one sheet per large valueset
     (FACILITY_TYPE, PAYER_TYPE, UNIT, ...). The SOURCES sheet maps each valueset
     to the CDM field(s) it governs. Stored once under `valuesets:` and referenced.

Output: valuesets/pcornet_cdm.yaml

    ./.venv/bin/pip install openpyxl pdfplumber pyyaml
    ./.venv/bin/python valuesets/build_catalog.py

Re-run when PCORnet releases a new spec or valueset workbook. Enumerable valuesets
ONLY — external code systems (ICD, LOINC, NDC, RxNorm, SNOMED) are not enumerable
and are intentionally excluded; validate those by code-system/format instead.
"""
import re
from pathlib import Path

import openpyxl
import pdfplumber
import yaml

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
PDF = ROOT / "docs" / "PCORI_CDM_SPECIFICATION.pdf"
XLSX = ROOT / "docs" / "PCORNET_ADDITIONAL_VALUESETS.xlsx"
OUT = HERE / "pcornet_cdm.yaml"

# Code lines come in two layouts: "P=Principal" (one token) and "Y = Diagnosis
# present" (spaced, e.g. DIAGNOSIS.DX_POA). Tolerate optional spaces around '='.
CODE_RE = re.compile(r"^([A-Za-z0-9_.+\-]{1,40})\s*=\s*(.+)$")


# ---------------------------------------------------------------- PDF inline sets
def _page_valuesets(pg, carry_table, anchors):
    """Return (table, {field: [(code, desc)]}, anchors) for one spec page.

    The valueset text sits in the x-band between the 'Predefined' and 'Definition'
    column headers. A table can span pages, and continuation pages do NOT repeat
    the header row — so when this page has its own headers we (re)compute the
    column x-anchors and remember them; otherwise we reuse the carried anchors so
    continuation rows (e.g. ENCOUNTER.ENC_TYPE, whose long definition spills onto
    a second page) are still captured."""
    words = pg.extract_words(keep_blank_chars=False)
    txt = pg.extract_text() or ""
    mt = re.search(r"^([A-Z][A-Z_ ]+?) Table Specification", txt, re.M)
    table = mt.group(1).strip().replace(" ", "_") if mt else carry_table
    if not words:
        return table, {}, anchors

    def xof(t):
        m = [w for w in words if w["text"] == t]
        return m[0]["x0"] if m else None

    if xof("Predefined") is not None:
        # real header page — recompute and remember column geometry
        hdr_top = [w for w in words if w["text"] == "Predefined"][0]["top"]
        anchors = {"xpred": xof("Predefined"), "xdef": xof("Definition"),
                   "fname_x": xof("RDBMS") or xof("Predefined"), "hdr_top": hdr_top}
    elif anchors:
        # continuation page — reuse geometry; rows start near the top
        anchors = {**anchors, "hdr_top": 30}
    else:
        return table, {}, anchors
    xpred, xdef, fname_x, hdr_top = (anchors["xpred"], anchors["xdef"],
                                     anchors["fname_x"], anchors["hdr_top"])
    if xpred is None or xdef is None:
        return table, {}, anchors

    # field names: left column caps tokens; merge fragments of a wrapped name
    fwords = [w for w in words if w["x0"] < fname_x - 2 and w["top"] > hdr_top + 8
              and re.match(r"^[A-Z][A-Z0-9_]*$", w["text"])]
    fwords.sort(key=lambda w: (round(w["top"]), w["x0"]))
    flist = []  # [top, name]
    for w in fwords:
        if flist and (w["top"] - flist[-1][0]) < 14:
            flist[-1][1] += w["text"]
        else:
            flist.append([w["top"], w["text"]])

    # valueset band -> visual lines
    band = [w for w in words if xpred - 3 <= w["x0"] < xdef - 3 and w["top"] > hdr_top + 8]
    band.sort(key=lambda w: (round(w["top"]), w["x0"]))
    lines, curtop, buf = [], None, []
    for w in band:
        if curtop is None or abs(w["top"] - curtop) < 4:
            buf.append(w); curtop = curtop or w["top"]
        else:
            lines.append((curtop, " ".join(x["text"] for x in buf))); buf = [w]; curtop = w["top"]
    if buf:
        lines.append((curtop, " ".join(x["text"] for x in buf)))

    def junk(s):
        return s.startswith(("http", "Source")) or "://" in s or "ViewValueSet" in s or "federalregister" in s

    entries = []  # [top, code, desc]
    for top, s in lines:
        s = s.strip()
        if not s or junk(s):
            continue
        m = CODE_RE.match(s)
        if m:
            entries.append([top, m.group(1), m.group(2).strip()])
        elif entries:
            entries[-1][2] = (entries[-1][2] + " " + s).strip()  # wrapped description

    flist.sort()
    out = {}
    for top, code, desc in entries:
        fld = None
        for ft, fn in flist:
            if ft <= top + 3:
                fld = fn
            else:
                break
        if fld:
            out.setdefault(fld, []).append((code, desc))
    return table, out, anchors


def extract_pdf(pdf):
    """{(table, field): [(code, desc)]} for every inline valueset, plus the CDM
    spec version, and a field->set(tables) index over ALL fields (for xlsx mapping)."""
    catalog, field_tables = {}, {}
    cur = None
    version = "unknown"
    anchors = None
    for i, pg in enumerate(pdf.pages):
        words = pg.extract_words(keep_blank_chars=False)
        txt = pg.extract_text() or ""
        if i == 0:
            mv = re.search(r"Specification,\s*Version\s*([\d.]+)", txt)
            if mv:
                version = mv.group(1)
        mt = re.search(r"^([A-Z][A-Z_ ]+?) Table Specification", txt, re.M)
        if mt:
            cur = mt.group(1).strip().replace(" ", "_")

        # field->tables index from the left column (every field, valueset or not)
        if cur:
            xr = [w for w in words if w["text"] == "RDBMS"]
            if xr:
                xrdbms, top0 = xr[0]["x0"], xr[0]["top"]
                frags = [w for w in words if w["x0"] < xrdbms - 2 and w["top"] > top0 + 8
                         and re.match(r"^[A-Z][A-Z0-9_]*$", w["text"])]
                frags.sort(key=lambda w: (round(w["top"]), w["x0"]))
                merged = []
                for w in frags:
                    if merged and (w["top"] - merged[-1][0]) < 14:
                        merged[-1][1] += w["text"]
                    else:
                        merged.append([w["top"], w["text"]])
                for _, name in merged:
                    field_tables.setdefault(name, set()).add(cur)

        _, vs, anchors = _page_valuesets(pg, cur, anchors)
        for fld, codes in vs.items():
            key = (cur, fld)
            catalog.setdefault(key, [])
            have = {c for c, _ in catalog[key]}
            for c, d in codes:
                if c not in have:
                    catalog[key].append((c, d)); have.add(c)
    return catalog, field_tables, version


# ------------------------------------------------------------- xlsx large sets
# The workbook (v1.14) predates CDM v7 field renames; map its SOURCES field
# labels to the actual v7 TABLE.COLUMN(s). Empty list => resolve via PDF index.
XLSX_FIELD_OVERRIDES = {
    "PATIENT_PREF_LANGUAGE_SPOKEN": ["DEMOGRAPHIC.PAT_PREF_LANGUAGE_SPOKEN",
                                     "PRIVATE_DEMOGRAPHIC.PAT_PREF_LANGUAGE_SPOKEN"],
    "PROVIDER_SPECIALTY_PRIMARY": ["PROVIDER.PROVIDER_SPECIALTY"],
}


def extract_xlsx(wb):
    """Return (version, {valueset_name: {"fields": [...], "values": {code: desc}}})."""
    info = wb["INFO"]
    rows = [r for r in info.iter_rows(values_only=True) if any(c is not None for c in r)]
    version = str(rows[-1][0]) if rows else "unknown"

    # SOURCES: valueset -> field name(s) (a cell may hold several, newline-joined)
    sources = {}
    for r in wb["SOURCES"].iter_rows(values_only=True):
        if r[0] and r[1] and r[0] != "VALUESET":
            fields = [f.strip() for f in str(r[1]).splitlines() if f.strip()]
            sources[str(r[0]).strip()] = fields

    valuesets = {}
    for name in wb.sheetnames:
        if name in ("INFO", "SOURCES"):
            continue
        ws = wb[name]
        values = {}
        header_seen = False
        for row in ws.iter_rows(values_only=True):
            if not header_seen:
                header_seen = True  # skip CODE/DESCRIPTIVE_TEXT header
                continue
            code = row[0]
            if code is None or str(code).strip() == "":
                continue
            desc = row[1] if len(row) > 1 and row[1] is not None else ""
            values[str(code).strip()] = str(desc).strip()
        valuesets[name] = {"fields": sources.get(name, []), "values": values}
    return version, valuesets


# ----------------------------------------------------------------------- build
def main():
    pdf = pdfplumber.open(PDF)
    inline, field_tables, cdm_version = extract_pdf(pdf)
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    xlsx_version, xlsx_sets = extract_xlsx(wb)

    columns = {}  # "TABLE.COLUMN" -> entry

    # 1. inline (PDF) valuesets — embedded values, self-contained per column
    for (table, field), codes in inline.items():
        if not table or not field:
            continue
        columns[f"{table}.{field}"] = {
            "kind": "enumerable",
            "source": "cdm_spec",
            "values": {c: d for c, d in codes},
        }

    # 2. xlsx valuesets — stored once under `valuesets:`, referenced by columns
    named = {}
    unresolved = []
    for name, vs in xlsx_sets.items():
        named[name] = {
            "kind": "enumerable",
            "source": f"additional_valuesets_v{xlsx_version}",
            "size": len(vs["values"]),
            "values": vs["values"],
        }
        targets = []
        for field in vs["fields"]:
            if field in XLSX_FIELD_OVERRIDES:
                targets += XLSX_FIELD_OVERRIDES[field]
            else:
                tabs = field_tables.get(field)
                if tabs:
                    targets += [f"{t}.{field}" for t in sorted(tabs)]
                else:
                    unresolved.append((name, field))
        for col in targets:
            columns[col] = {"kind": "enumerable", "source": f"additional_valuesets_v{xlsx_version}",
                            "valueset": name}

    catalog = {
        "meta": {
            "cdm": "pcornet",
            "cdm_spec_version": cdm_version,
            "additional_valuesets_version": xlsx_version,
            "generated_from": ["docs/PCORI_CDM_SPECIFICATION.pdf",
                               "docs/PCORNET_ADDITIONAL_VALUESETS.xlsx"],
            "scope": ("Enumerable valuesets only. External code systems (ICD, LOINC, "
                      "NDC, RxNorm, SNOMED) are excluded — validate those by code "
                      "system/format, not membership."),
            "regenerate": "valuesets/build_catalog.py",
        },
        "valuesets": dict(sorted(named.items())),
        "columns": dict(sorted(columns.items())),
    }

    OUT.write_text(
        "# PCORnet CDM valueset catalog — GENERATED by valuesets/build_catalog.py.\n"
        "# Do not edit by hand; re-run the generator against the source docs.\n"
        + yaml.safe_dump(catalog, sort_keys=False, allow_unicode=True, width=100)
    )

    n_inline = sum(1 for c in columns.values() if "values" in c)
    n_ref = sum(1 for c in columns.values() if "valueset" in c)
    print(f"CDM spec v{cdm_version} + additional valuesets v{xlsx_version}")
    print(f"wrote {OUT.relative_to(ROOT)}")
    print(f"  {len(columns)} constrained columns ({n_inline} inline, {n_ref} reference)")
    print(f"  {len(named)} large named valuesets: {', '.join(sorted(named))}")
    if unresolved:
        print(f"  WARNING unresolved xlsx field->table ({len(unresolved)}): {unresolved}")


if __name__ == "__main__":
    main()
